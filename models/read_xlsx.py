import os

import openpyxl
import openpyxl_image_loader

from base64 import b64encode
from base64 import b64decode
from io import BytesIO

from PIL import Image


def binary_photo(pict: str) -> bytes:
    with open(pict, 'rb') as picture:
        return b64encode(picture.read())


def read_binary_photo(binary: str) -> Image:
    image = BytesIO(b64decode(binary))
    image = Image.open(image)
    # image.show()
    return image


def photo_change(image: Image):
    fixed_width = 200
    width_percent = (fixed_width / float(image.size[0]))
    height_size = int((float(image.size[0]) * float(width_percent)))
    image = image.resize((fixed_width, height_size))
    image.save(fp='image', format='PNG')
    image = binary_photo(pict='image')
    image = str(image, 'utf-8')
    os.remove(path='image')
    return image


def xlsx_to_dict(path: str) -> dict:
    """
    Function open a xlsx file and write its content down to a dictionary.
    Don't write headlines of column (manually use row=2 of file)

    :param path: absolute path to xlsx file
    :type path: str

    :return: dictionary with xlsx file content
    :rtype: dict
    """
    data = {}
    try:
        book = openpyxl.load_workbook(path)
    except Exception:
        return None
    worksheet = book.active
    image_loader = openpyxl_image_loader.SheetImageLoader(worksheet)

    for row in range(2, worksheet.max_row + 1):
        data[str(worksheet[row][worksheet.min_column - 1].value)] = []
        for col in range(1, worksheet.max_column):
            index = str(worksheet.cell(row, col + 1)).split('.')[1].strip('>')
            if image_loader.image_in(index):
                try:
                    image = image_loader.get(index)
                    image = photo_change(image=image)
                    data[str(worksheet[row][worksheet.min_column - 1].value)].append(image)
                except ValueError:
                    data[str(worksheet[row][worksheet.min_column - 1].value)].append('None')
            else:
                data[str(worksheet[row][worksheet.min_column - 1].value)].append(str(worksheet[row][col].value))
    return data


if __name__ == '__main__':
    # path = '/home/mikhail/Рабочий стол/Coding/my_projects/xlsx_to_postgres/samples/временные пропуска персонала.xlsx'
    # path = '/home/mikhail/Рабочий стол/Coding/my_projects/xlsx_to_postgres/samples/временные пропуска ТС.xlsx'
    # path = '/home/mikhail/Рабочий стол/Coding/my_projects/xlsx_to_postgres/samples/по спискам персонал.xlsx'
    # path = '/home/mikhail/Рабочий стол/Coding/my_projects/xlsx_to_postgres/samples/по спискам ТС.xlsx'
    path = '/home/mikhail/Рабочий стол/Coding/my_projects/xlsx_to_postgres/samples/ТТН.xlsx'
    data = xlsx_to_dict(path=path)
    print(data)
